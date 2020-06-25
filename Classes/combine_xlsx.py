import pandas as pd
import glob
from openpyxl import Workbook
from General.Functions import *
from openpyxl.utils import get_column_letter

results_folder = "Results"

master_workbook = Workbook()
master_workbook.remove_sheet(master_workbook.active)

excel_names = list(glob.glob(results_folder + "/*.xlsx"))
print(excel_names)

for excel_file in excel_names:
    wb = open_xlsx(excel_file)
    ws = wb.active

    print("Appending " + ws.title)
    master_workbook.create_sheet(ws.title)
    sheet = master_workbook[ws.title]
    for r_idx, row in enumerate(ws.iter_rows()):
        for c_idx, cell in enumerate(row):
            sheet[get_column_letter(c_idx+1) + str(r_idx+1)] = cell.value

print("Saving")
count = 0
try:
    while True:
        master_workbook.save(results_folder + "/" + "Master.xlsx")
        break
except:
    if count == 0:
        print("Close Excel Workbook!")
    count = 1
