import subprocess
import sys
import glob
import dateutil
from Stock import *
from Hurricane import *
from openpyxl import Workbook
import itertools
# getting Functions.py from Stocks.py


def main():
    wiki_sap_script = "wiki_sap.R"
    get_hist_stock_prices_script = "get_hist_stock_prices.R"
    wiki_sap_name = "wiki_sap.txt"
    temp_wiki_sap_name = "temp_" + wiki_sap_name
    ticker_list_txt_name = "tickers.txt"
    stock_data_name = "stock_data.xlsx"
    hurricane_data_folder = "Hurricane Data"
    results_folder = "Results"
    rscript_path = "C:/Program Files/R/R-3.5.1/bin/Rscript.exe"

    master_gics_filters = ["Health Care", "Real Estate", "Industrials"] #, "Real Estate", "Industrials", "Energy", "Utilities", "Utilities", "Financials", "Communication Services"
    master_gics_filters_combinations = []
    for L in range(0, len(master_gics_filters) + 1):
        for subset in itertools.combinations(master_gics_filters, L):
            subset = list(subset)
            if subset:
                print(subset)
                master_gics_filters_combinations.append(subset)
        print("")

    stock_start_year = 2008
    hurricane_list = get_hurricane_list(hurricane_data_folder, stock_start_year)

    for master_gics_filters_combination in master_gics_filters_combinations:

        sheet_title = ""
        for i, element in enumerate(master_gics_filters_combination):
            if i != 0:
                sheet_title += "."
            for letter in element:
                if letter in "QWERTYUIOPASDFGHJKLZXCVBNM":
                    sheet_title += letter

        if results_folder + "\\" + sheet_title + ".xlsx" in list(glob.glob(results_folder + "/*.xlsx")):
            print(sheet_title + " already ran, running next combination")
            continue
        else:
            print("Running " + sheet_title)

        stock_gics_filter_list = master_gics_filters_combination
        stock_gics_sub_filter_list = []
        plus_min = "plus"
        stock_list_limit = 5
        stock_start_year = 2008

        clean_folder(wiki_sap_name, stock_data_name, temp_wiki_sap_name)

        stock_list = get_stock_list(hurricane_list, rscript_path, wiki_sap_script, stock_list_limit, get_hist_stock_prices_script, wiki_sap_name, stock_data_name, ticker_list_txt_name, stock_start_year, temp_wiki_sap_name, stock_gics_filter_list, stock_gics_sub_filter_list, plus_min)

        wb = Workbook()
        wb.remove_sheet(wb.active)
        wb.create_sheet(sheet_title)
        sheet = wb[sheet_title]
        row = 1
        col = 1
        col_name_list = ["Date", "Ticker", "Co. Loc", "Hurricane", "Distance (mi)", "% change", "Windspeed", "Category"]

        for col_name in col_name_list:
            sheet.cell(row=row, column=col).value = col_name
            col += 1
        row += 1
        col = 1
        loop_cond = True
        for stock in stock_list:
            for key in stock.hurricane_dict.keys():
                for date_dict in stock.hurricane_dict[key]:
                    print("ROW: " + str(row))
                    if row == 1048575:
                        loop_cond = False
                        break

                    date_data_list = stock.hurricane_dict[key][date_dict]
                    for date_data in date_data_list:
                        sheet.cell(row=row, column=col).value = date_data
                        col += 1
                    row += 1
                    col = 1
                if not loop_cond:
                    break
            if not loop_cond:
                break

        count = -10
        while True:
            try:
                wb.save(results_folder + "/" + sheet_title + '.xlsx')
                break
            except:
                if count >= 10:
                    print("Close Excel Workbook!")
                count += 1


def clean_folder(wiki_sap_name, stock_data_name, temp_wiki_sap_name):
    remove_file(wiki_sap_name)
    remove_file(stock_data_name)
    remove_file(temp_wiki_sap_name)


def filter_stock_list(stock_list, start_year=None):
    print("\tFiltering stock dates")
    new_stock_list = []
    for stock in stock_list:
        new_price_dict = {}
        if start_year is not None:
            for key in stock.price_dict.keys():
                if key.year >= start_year:
                    new_price_dict[key] = stock.price_dict[key]
        stock.price_dict = new_price_dict
        new_stock_list.append(stock)
    return new_stock_list


def get_stock_list(hurricane_list, rscript_path, wiki_sap_script, stock_list_limit, get_hist_stock_prices_script, wiki_sap_name, stock_data_name, ticker_list_txt_name, stock_start_year, temp_wiki_sap_name, stock_gics_filter_list, stock_gics_sub_filter_list, plus_min):

    subprocess.call([rscript_path, "--vanilla", wiki_sap_script])
    wiki_sap_filter_gics(wiki_sap_name, temp_wiki_sap_name, stock_gics_filter_list, stock_gics_sub_filter_list, plus_min)
    wiki_sap_txt_to_tickers_txt(wiki_sap_name, ticker_list_txt_name, end=stock_list_limit)
    subprocess.call([rscript_path, "--vanilla", get_hist_stock_prices_script])

    print("\n" + "-" * 60)
    print("\nGetting stocks...")

    wiki_sap_txt = open(wiki_sap_name, "r")
    ticker_txt_start_idx = 0
    stock_list = []
    wiki_sap_xlsx = open_xlsx(stock_data_name)
    for count, i in enumerate(wiki_sap_txt.readlines()):

        if i.strip().rstrip().split(";")[0].lstrip().rstrip() not in txt_to_list(ticker_list_txt_name):

            ticker_txt_start_idx += stock_list_limit
            print("Running get_hist_stock_prices_script.R subprocess")
            wiki_sap_txt_to_tickers_txt(wiki_sap_name, ticker_list_txt_name, start=ticker_txt_start_idx, end=ticker_txt_start_idx + stock_list_limit)
            print("\tFetching - " + str(txt_to_list(ticker_list_txt_name)))
            subprocess.call([rscript_path, "--vanilla", get_hist_stock_prices_script], stdout=open(os.devnull, 'w'),
                            stderr=subprocess.STDOUT)
            continue

        temp_stock = Stock(i.strip().rstrip(), hurricane_list, wiki_sap_xlsx, stock_start_year)

        if not temp_stock.coordinates or temp_stock.price_dict == {}:
            if not temp_stock.coordinates:
                print(temp_stock.tic + " was removed - no coordinate could be found " + temp_stock.loc)
            elif temp_stock.price_dict == {}:
                print(temp_stock.tic + " was removed - no historical stock data, check IPO vs start_year filter")

            continue

        print(string_length(str(len(stock_list) + 1) + ") ", 5) + str(temp_stock))
        stock_list.append(temp_stock)

        if stock_list_limit is not None:
            if len(stock_list) >= stock_list_limit:
                break
    wiki_sap_txt.close()

    if not stock_list:
        print("Stock list is empty. Check conditionals!")
        exit()
    print("-" * 60 + "\n")
    stock_list = filter_stock_list(stock_list, start_year=stock_start_year)

    return stock_list


def filter_hurricane_list(hurricane_list, start_year=None):
    print("\n\tFiltering hurricane dates")
    new_hurricane_list = []
    for hurricane in hurricane_list:
        if start_year is not None:
            if list(hurricane.data_dict.keys())[0].year >= start_year:
                new_hurricane_list.append(hurricane)
    return new_hurricane_list


def get_hurricane_list(hurricane_data_folder, stock_start_year):
    print("\n" + "-" * 60)
    print("\nParsing hurricane data to Hurricane objects")
    txt_list_list = []
    for txt_filepath in glob.glob(hurricane_data_folder + "/*.txt"):
        txt_file = open(txt_filepath, "r")
        for line in txt_file.readlines():
            txt_list = []
            for i, part in enumerate(line.split(',')):
                if i >= 8:
                    break
                if part != "\n" or not part.isspace():
                    if i == 0 and len(line.split(',')) > 4:
                        txt_list.append(dateutil.parser.parse(part.strip().rstrip()))
                    else:
                        txt_list.append(part.strip().rstrip())
            txt_list_list.append(txt_list)
        txt_file.close()

    hurricane_list = []
    dict_entry = ""
    hurricane_raw_list = []
    for i, data_list in enumerate(txt_list_list):
        if len(data_list) < 4:
            if dict_entry == "":
                dict_entry = data_list
            else:
                hurricane_list.append(Hurricane(dict_entry, hurricane_raw_list))
                sys.stdout.write("\r\t{0}) Hurricane: {1}: {2}".format(i + 1, dict_entry, string_length(str(hurricane_raw_list), 20, dots=True)))
                sys.stdout.flush()
                dict_entry = data_list
                hurricane_raw_list = []
        else:
            hurricane_raw_list.append(data_list)

    hurricane_list = filter_hurricane_list(hurricane_list, stock_start_year)
    print("")
    print("-" * 60 + "\n")
    return hurricane_list


def wiki_sap_txt_to_tickers_txt(wiki_sap_name, ticker_list_txt_name, start=0, end=None):
    written_counter = 0
    wiki_sap_txt = open(wiki_sap_name, "r")
    tickers_txt = open(ticker_list_txt_name, 'w')
    for i, stock_line in enumerate(wiki_sap_txt.readlines()):
        if start > i:
            continue
        if written_counter != 0:
            tickers_txt.write("\n")

        tickers_txt.write(stock_line.split(";")[0].lstrip().rstrip())
        written_counter += 1
        if end is not None:
            if end == i + 1:
                break
    wiki_sap_txt.close()
    tickers_txt.close()


def wiki_sap_filter_gics(wiki_sap_name, temp_wiki_sap_name, stock_gics_filter_list, stock_gics_sub_filter_list, plus_min):
    written_counter = 0
    wiki_sap_txt = open(wiki_sap_name, "r")
    new_wiki_sap_txt = open(temp_wiki_sap_name, 'w')
    for i, stock_line in enumerate(wiki_sap_txt.readlines()):
        gics = stock_line.split(";")[2].lstrip().rstrip()
        gics_sub = stock_line.split(";")[3].lstrip().rstrip()

        if stock_gics_filter_list or stock_gics_sub_filter_list:
            if plus_min == "plus":
                if gics in stock_gics_filter_list or gics_sub in stock_gics_sub_filter_list:
                    if written_counter != 0:
                        new_wiki_sap_txt.write("\n")
                    new_wiki_sap_txt.write(stock_line.lstrip().rstrip())
                    written_counter += 1
            else:
                if not (gics in stock_gics_filter_list) and not (gics_sub in stock_gics_sub_filter_list):
                    new_wiki_sap_txt.write(stock_line.lstrip().rstrip())
                    written_counter += 1
                    if written_counter != 0:
                        new_wiki_sap_txt.write("\n")
        else:
            new_wiki_sap_txt.write(stock_line.lstrip().rstrip())
            written_counter += 1
            if written_counter != 0:
                new_wiki_sap_txt.write("\n")

    wiki_sap_txt.close()
    new_wiki_sap_txt.close()

    new_wiki_sap_txt = open("temp_" + wiki_sap_name, 'r')
    wiki_sap_txt = open(wiki_sap_name, "w")
    for line in new_wiki_sap_txt.readlines():
        wiki_sap_txt.write(line)
    wiki_sap_txt.close()
    new_wiki_sap_txt.close()


def to_excel(stock_list):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Testing"
    row = 1
    col = 1
    col_name_list = ["Date", "Ticker", "Hurricane", "Distance (mi)", "% change", "Windspeed", "Category"]
    for col_name in col_name_list:
        sheet.cell(row=row, column=col).value = col_name
        col += 1
    row += 1
    col = 1
    for stock in stock_list:
        for key in stock.hurricane_dict.keys():
            for date_dict in stock.hurricane_dict[key]:
                date_data_list = stock.hurricane_dict[key][date_dict]
                for date_data in date_data_list:
                    sheet.cell(row=row, column=col).value = date_data
                    col += 1
                row += 1
                col = 1

    count = 0
    while True:
        try:
            wb.save('Regression Data.xlsx')
            break
        except:
            if count == 0:
                print("Close Excel Workbook!")
            count = 1

main()
